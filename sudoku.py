# Working code to Generate a puzzle
# -*- coding: utf-8 -*-
import copy
import math
import os
import random
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Difficulty level can be 'Easy' | 'Medium' | 'Hard' | 'Insane'
level = "Medium"
size = 3
amount = 2  # Number of Sudoku to produce
print_console = False  # Print to Python console
export_excel = True  # Export Sudoku to excel


# Sudoku sizes, size is grid size.
# size -> x * y = number of cells -> using digits 1 ... #
# 2 -> 2 * 2 = 4 -> numbers 1 ... 4
# 3 -> 3 * 3 = 9 -> numbers 1 ... 9
# 4 -> 4 * 4 = 16 -> numbers 1 ... 16
# 5 -> 5 * 5 = 25 -> numbers 1 ... 25
# ...

# Number of Sudoku per page given the size
# size -> Sudoku per page
# 2 -> 18 per page
# 3 -> 2 per page
# 4 -> 1 per page
# ...


# Initializes cell object.
# A cell is a single box of a Sudoku puzzle.
# 81 cells make up the body of a 3*3 puzzle.
class Cell:
    def __init__(self, position):
        self.notes = list(range(1, size ** 2 + 1)) # Notes are the possible answers for a Cell
        self.answer = None
        self.position = position
        self.solved = False

    # Remove a digit from the list of possible answers
    # Mark cell solved if one digit remains
    def remove_digit_from_notes(self, num):
        if num in self.notes and self.solved is False:
            self.notes.remove(num)
            if len(self.notes) == 1:
                self.answer = self.notes[0]
                self.solved = True
        if num in self.notes and self.solved is True:
            self.answer = 0

    # Return solved status of a Cell
    def is_cell_sovled(self):
        return self.solved

    # Return the position of cell within a Sudoku puzzle x = row; y = col; z = box number
    def get_cell_position(self):
        return self.position

    # Return possible answers for cell
    def return_possible(self):
        return self.notes

    # Return length of possible answers for cell
    def len_of_possible(self):
        return len(self.notes)

    # Return answer for cell if solved, return 0 otherwise
    def return_answer_if_solved(self):
        if self.solved:
            return self.notes[0]
        else:
            return 0

    # Hard code a digit, clean out the possible digits, and set solved to True
    def set_answer(self, num):
        if num in list(range(1, size ** 2 + 1)):
            self.solved = True
            self.answer = num
            self.notes = [num]
        else:
            raise (ValueError)

    # Reset the possible digits to the default, set solved to False, remove answer if found
    def reset_cell(self):
        self.notes = list(range(1, size ** 2 + 1))
        self.answer = None
        self.solved = False


# Create an empty list of empty Cells, grid[0] is Cell.position = 1
def empty_sudoku():
    grid = []
    for x in list(range(1, size ** 2 + 1)):
        for y in list(range(1, size ** 2 + 1)):
            z = ((x - 1) * (size ** 2)) + y
            c = Cell((x, y, z))
            grid.append(c)
    return grid


# Return a difficulty level depending on the number of guesses required to solve
def decide_difficulty_level(guesses):
    if guesses == 0:
        return 'Easy'
    elif guesses <= 2:
        return 'Medium'
    elif guesses <= 7:
        return 'Hard'
    return 'Insane'


# Print sudoku to terminal
def print_sudoku(sudoku):
    width = len(str(size ** 2))
    columns = size ** 2

    for row in range(columns * 2 + 1):
        print("")
        for column in range(columns * 4 + 1):
            if row == 0:
                if column == 0:
                    print("╔", end='')
                elif column % 4 == 0:
                    if column == columns * 4:
                        print("╗", end='')
                    elif column % (4 * size) == 0:
                        print("╦", end='')
                    else:
                        print("╤", end='')
                elif (column + 2) % 4 == 0:
                    print("═" * width, end='')
                elif column % 4 > 0:
                    print("═", end='')
            elif row % 2 > 0:
                if column % (4 * size) == 0:
                    print("║", end='')
                elif column % 4 == 0:
                    print("│", end='')
                elif (column + 2) % 4 == 0:
                    i = int((row - 1) / 2 * columns + (column - 1) / 4)
                    value = sudoku[i].return_answer_if_solved()
                    if value == 0:
                        value = ' '
                    print(f"{str(value).rjust(width)}", end='')
                else:
                    print(' ', end='')
            elif row % (columns * 2) == 0:
                if column == 0:
                    print("╚", end='')
                elif column % 4 == 0:
                    if column == columns * 4:
                        print("╝", end='')
                    elif column % (4 * size) == 0:
                        print("╩", end='')
                    else:
                        print("╧", end='')
                elif (column + 2) % 4 == 0:
                    print("═" * width, end='')
                elif column % 4 > 0:
                    print("═", end='')
            elif row % (size * 2) == 0:
                if column == 0:
                    print("╠", end='')
                elif column % 4 == 0:
                    if column == columns * 4:
                        print("╣", end='')
                    elif column % (4 * size) == 0:
                        print("╬", end='')
                    else:
                        print("╪", end='')
                elif (column + 2) % 4 == 0:
                    print("═" * width, end='')
                elif column % 4 > 0:
                    print("═", end='')
            else:
                if column == 0:
                    print("╟", end='')
                elif column % 4 == 0:
                    if column == columns * 4:
                        print("╢", end='')
                    elif column % (4 * size) == 0:
                        print("╫", end='')
                    else:
                        print("┼", end='')
                elif (column + 2) % 4 == 0:
                    print("─" * width, end='')
                elif column % 4 > 0:
                    print("─", end='')
    print("")


# Export sudoku to Excel
def export_sudoku_to_excel(sudoku, count=0, solution=False):
    columns = size ** 2

    if solution:
        filename = f'sudoku_{columns}x{columns}_{level}_Solution.xlsx'
    else:
        filename = f'sudoku_{columns}x{columns}_{level}.xlsx'

    MAX_COLUMN_PER_PAGE = 16
    MAX_ROW_PER_PAGE = 28

    max_sudoku_per_row = max(1, math.floor(MAX_COLUMN_PER_PAGE / (columns + 1)))

    position_column = count % max_sudoku_per_row
    position_row = math.floor(count / max_sudoku_per_row)

    offset_column = position_column * (columns + 1)
    offset_row = position_row * (columns + 1)

    if os.path.isfile(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.active

    thin = Side(border_style="thin", color="00808080")
    thick = Side(border_style="thick", color="00808080")

    border = Border(left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin,
                    )
    borderlefttop = Border(left=thick,
                           right=thin,
                           top=thick,
                           bottom=thin,
                           )
    borderleft = Border(left=thick,
                        right=thin,
                        top=thin,
                        bottom=thin,
                        )
    borderleftbottom = Border(left=thick,
                              right=thin,
                              top=thin,
                              bottom=thick,
                              )
    borderbottom = Border(left=thin,
                          right=thin,
                          top=thin,
                          bottom=thick,
                          )
    borderrightbottom = Border(left=thin,
                               right=thick,
                               top=thin,
                               bottom=thick,
                               )
    borderright = Border(left=thin,
                         right=thick,
                         top=thin,
                         bottom=thin,
                         )
    borderrighttop = Border(left=thin,
                            right=thick,
                            top=thick,
                            bottom=thin,
                            )
    bordertop = Border(left=thin,
                       right=thin,
                       top=thick,
                       bottom=thin,
                       )

    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    for row in range(columns):
        ws.row_dimensions[row + 1 + offset_row].height = 25
        for column in range(columns):
            ws.column_dimensions[get_column_letter(column + 1 + offset_column)].width = 5
            value = sudoku[row * columns + column].return_answer_if_solved()
            if value == 0:
                value = ''
            cell = ws.cell(row=row + 1 + offset_row, column=column + 1 + offset_column, value=value)
            if row % size == 0 and column % size == 0:
                cell.border = borderlefttop
            elif row % size == 0 and (column + 1) % size == 0:
                cell.border = borderrighttop
            elif row % size == 0:
                cell.border = bordertop
            elif (row + 1) % size == 0 and column % size == 0:
                cell.border = borderleftbottom
            elif (row + 1) % size == 0 and (column + 1) % size == 0:
                cell.border = borderrightbottom
            elif (row + 1) % size == 0:
                cell.border = borderbottom
            elif column % size == 0:
                cell.border = borderleft
            elif (column + 1) % size == 0:
                cell.border = borderright
            else:
                cell.border = border
            cell.alignment = alignment

    wb.save(filename)


def print_sudoku_sequence(sudoku):
    cell_values = []
    for c in range(0, len(sudoku)):
        cell_values.append(sudoku[c].return_answer_if_solved())

    print(cell_values)
    return cell_values


# Generate a random Sudoku
def sudoku_generator():
    # Generate list of cell positions, for 3x3 grid that is 81
    cells = [i for i in range(size ** 4)]

    # Generate a list of empty Cells
    sudoku = empty_sudoku()

    while cells:
        # For each remaining Cell check how many digits it could still contain
        # Return the number of digits possible
        lowest_number_of_notes = []
        for c in cells:
            lowest_number_of_notes.append(
                sudoku[c].len_of_possible())
        m = min(lowest_number_of_notes)

        # Get all Cells with the same number of digits possible
        lowest = []
        for c in cells:
            if sudoku[c].len_of_possible() == m:
                lowest.append(sudoku[c])

        # Randomly choose one of the Cells
        random_cell = random.choice(lowest)
        # Get the index of that Cell in the sudoku
        random_cell_index = sudoku.index(random_cell)
        # Remove the Cell from the cells list
        cells.remove(random_cell_index)

        # Get the cell's value, if 0 guess one from its notes
        random_cell_value = random_cell.return_answer_if_solved()
        if random_cell_value == 0:  # the actual setting of the cell
            possible_values = random_cell.return_possible()

            # Guess a random value for this random Cell
            random_cell_value = random.choice(possible_values)
            random_cell.set_answer(random_cell_value)

        # Remove the cell digit from all neighbours
        position1 = random_cell.get_cell_position()
        for c in cells:
            position2 = sudoku[c].get_cell_position()
            if position1[0] == position2[0]:
                sudoku[c].remove_digit_from_notes(random_cell_value)
            if position1[1] == position2[1]:
                sudoku[c].remove_digit_from_notes(random_cell_value)
            if position1[2] == position2[2]:
                sudoku[c].remove_digit_from_notes(random_cell_value)
    return sudoku


# Test to see if the given Sudoku puzzle is valid
# Tests rows, columns and cells
# Returns True if valid, False if not
def sudoku_checker(sudoku):
    for i in range(len(sudoku)):
        for n in range(len(sudoku)):
            if i != n:
                position1 = sudoku[i].get_cell_position()
                position2 = sudoku[n].get_cell_position()
                if position1[0] == position2[0] or position1[1] == position2[1] or position1[2] == position2[2]:
                    num1 = sudoku[i].return_answer_if_solved()
                    num2 = sudoku[n].return_answer_if_solved()
                    if num1 == num2:
                        return False
    return True


# Generate a completed Sudoku
# Generated Sudoku is valid & randomized
def gen_completed_sudoku():
    sudoku = sudoku_generator()
    result = sudoku_checker(sudoku)

    # Only return the Sudoku if it is a valid Sudoku
    while not result:
        sudoku = sudoku_generator()
        result = sudoku_checker(sudoku)

    return sudoku


# Attempt to solve the Sudoku - uses Brute Force solving
# First find all Cells with just one possible digit and update all rows, columns and blocks
# Loop over the new grid and try to do the same thing again
# Once no single possible digit exists for a Cell, start guessing digits and see if the puzzle still solves
# Return the last solvable puzzle and the numbers of guesses it took
# Number of guesses will be an indication of the difficulty
def solver(sudoku, f=0):
    # Attempt a max of 900 solves
    if f > 900:
        print('Too many solve attempts')
        return False
    guesses = 0

    # Create a deep copy of the Sudoku as not to change the given puzzle
    copy_s = copy.deepcopy(sudoku)

    # Create a list of cell positions
    cells = [i for i in range(size ** 4)]

    # Create a list of solved Cells,
    # The list will be used to clean the notes in the same row, column and block
    solved_cells = []
    for c in cells:
        if copy_s[c].len_of_possible() == 1:
            solved_cells.append(c)

    # Iterate through the solved Cells and clean up the notes in the neighbours
    # If a neighbour has one possible note left we add it to the solved Cells list
    while solved_cells:
        for n in solved_cells:
            cell_position = copy_s[n].get_cell_position()

            # Get the value of the Cell
            cell_answer = copy_s[n].return_answer_if_solved()

            # Iterate through the remaining unset cells and remove digit from the Cell notes
            for c in cells:
                position2 = copy_s[c].get_cell_position()
                if cell_position[0] == position2[0]:
                    copy_s[c].remove_digit_from_notes(cell_answer)
                if cell_position[1] == position2[1]:
                    copy_s[c].remove_digit_from_notes(cell_answer)
                if cell_position[2] == position2[2]:
                    copy_s[c].remove_digit_from_notes(cell_answer)

                # Add Cell to solved list if the Cell is solved after deleting the current digit
                if copy_s[c].len_of_possible() == 1 and c not in solved_cells:
                    solved_cells.append(c)

            # Remove Cell from lists, so it won't be checked again
            solved_cells.remove(n)
            if n in cells:
                cells.remove(n)

        # If the list of Cells that still need to be solved is not empty but there are no solved_cells left to check
        # Start brute forcing the solution by randomly testing digits
        if cells != [] and solved_cells == []:
            # Start by finding the lowest number of notes left
            number_of_notes = []
            for c in cells:
                number_of_notes.append(copy_s[c].len_of_possible())
            m = min(number_of_notes)

            # Grab all cells that have this number of notes
            lowest = []
            for i in cells:
                if copy_s[i].len_of_possible() == m:
                    lowest.append(copy_s[i])

            # Randomly pick one of the Cells
            random_cell = random.choice(lowest)
            # Get the index of the Cell in the cells list
            cell_index = copy_s.index(random_cell)
            # Choose a random digit from the notes of the Cell
            random_digit = random.choice(copy_s[cell_index].return_possible())
            # Set the value of the Cell to the randomly selected digit
            copy_s[cell_index].set_answer(random_digit)
            # Now add the Cell to the solved list and attempt to solve other Cells
            solved_cells.append(random_digit)
            # Increment our guesses to increase the difficulty rating
            guesses += 1

    # Test to see if a valid solution was found and if so how many guesses it took
    if sudoku_checker(copy_s):
        print(f'Solved sudoku in {guesses} guesses')
        return copy_s, guesses
    # If not attempt to solve again
    else:
        return solver(sudoku, f + 1)


# Start solving a Sudoku using the solver
# Returns [Sudoku Puzzle, number of guesses] or False
def solve(sudoku):
    s = solver(sudoku)
    if s:
        return s

    return False


# Empty the Cells of a completed sudoku one by one, test the Sudoku,
# Returns [Sudoku Puzzle, number of guesses] or False
def generate_sudoku(sudoku):
    # Create list of cell positions, for 3*3 this is 81 Cell positions
    cells = [i for i in range(size ** 4)]
    cells_emptied = 0

    while cells:
        # Create a deepcopy of the sudoku
        copy_s = copy.deepcopy(sudoku)

        # Select a position to empty
        rand_index = random.choice(cells)
        # Remove position from the list
        cells.remove(rand_index)

        print(f"Removing {rand_index}")

        # Empty Cell and reset all settings on it
        copy_s[rand_index].reset_cell()

        # Attempt to solve the Sudoku after removing a digit from a Cell
        solved = solve(copy_s)

        # If after removing this digit the Sudoku can no longer be solved
        # Return the Sudoku in the state before this loop
        if solved[0]:
            # If the Sudoku can still be solved remove the same empty cell from the original Sudoku
            # Make sure the solution results in the same board state if solved again
            single_answer = False
            if equal_checker(solved[0], solve(copy_s)[0]):
                # Run the check again to make sure there is no second solution to the puzzle
                if equal_checker(solved[0], solve(copy_s)[0]):
                    print(f'Emptying Cell {rand_index}')
                    sudoku[rand_index].reset_cell()
                    cells_emptied += 1
                    single_answer = True
                    # print(sudoku[rand_index].return_answer_if_solved())
            if not single_answer:
                # If the Sudoku cannot be solved and there is more than 1 answer possible for the new board return the
                # Sudoku as it was in the previous loop
                print(f'Found more than 1 answer, returning sudoku after emptying {cells_emptied} cells')
                print_sudoku_sequence(sudoku)
                f = solve(sudoku)
                return f
        else:
            print(f'Could no longer solve the sudoku, returning sudoku after emptying {cells_emptied} cells')
            print_sudoku_sequence(sudoku)
            f = solve(sudoku)
            print("Guesses: " + str(f[1]))
            print("Level: " + decide_difficulty_level(f[1]))
            return f

# Test two board results to make sure they're equal
def equal_checker(s1, s2):
    for i in range(len(s1)):
        if s1[i].return_answer_if_solved() != s2[i].return_answer_if_solved():
            return False
    return True


# Generate a single puzzle, use the global level and size settings
# Numbers the puzzle with the count value
def main():
    # Get time at start of creation
    t1 = time.time()
    puzzles_created = []

    # while len(puzzles_created) < amount:
    print("------------------------------")
    print(f"Sudoku number: {len(puzzles_created) + 1}")

    # Get a completed & valid Sudoku
    # completed meaning all Cells have a valid digit
    completed_puzzle = gen_completed_sudoku()

    puzzle_attempt = 0
    max_attempts = 100

    while puzzle_attempt < max_attempts:
        # Use the completed puzzle and empty cells to create a Sudoku
        sudoku = generate_sudoku(completed_puzzle)
        sudoku_level = decide_difficulty_level(sudoku[1])

        # If the Sudoku is of the desired level add it to our list and create the next one
        if sudoku_level == level:
            puzzles_created.append(sudoku)
            print(f"{puzzle_attempt+1} created a valid puzzle")
            print_sudoku_sequence(sudoku[0])
            break

        puzzle_attempt += 1

    # Get time after creating puzzles
    t2 = time.time()
    t3 = t2 - t1
    print("Runtime is " + str(t3) + " seconds")

    print(f'Created {len(puzzles_created)} puzzles at difficulty level {level} in {puzzle_attempt} attempts')

    # if export_excel:
    #     export_sudoku_to_excel(sudoku[0], count)
    # if print_console:
    #     print_sudoku(sudoku[0])
    return


main()
